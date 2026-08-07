"""
report_generator.py — Gerador de Relatório Final para o TCC "Cedro Plus".

Lê os arquivos gerados (casos de teste preenchidos + defeitos) e calcula as
métricas exigidas na Entrega 4 do PDF:
  - Quantidade de testes realizados
  - Quantidade de testes aprovados
  - Quantidade de testes reprovados
  - Erros encontrados / corrigidos / permanecentes
  - Funcionalidades não testadas

Gera um relatório em Markdown com as estatísticas e uma conclusão final.

Uso:
    python report_generator.py [--planilha PATH] [--saida PATH]
"""

import argparse
import json
import os
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook

import config


def _carregar_casos(planilha: str | None = None) -> list:
    """Carrega os casos de teste da planilha Excel."""
    if planilha is None:
        planilha = os.path.join(config.DADOS_DIR, "casos_de_teste.xlsx")
    if not os.path.exists(planilha):
        print(f"[AVISO] Planilha não encontrada: {planilha}")
        return []
    wb = load_workbook(planilha, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    casos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        casos.append(dict(zip(headers, row)))
    return casos


def _carregar_defeitos() -> list:
    """Carrega os defeitos do arquivo JSON."""
    caminho = os.path.join(config.DADOS_DIR, "defeitos.json")
    if not os.path.exists(caminho):
        return []
    with open(caminho, "r", encoding="utf-8") as f:
        return json.load(f)


def _carregar_resultado_pytest() -> dict | None:
    """Carrega o resultado mais recente do pytest (relatório JSON)."""
    relatorios = [f for f in os.listdir(config.RELATORIOS_DIR)
                  if f.startswith("report_") and f.endswith(".json")]
    if not relatorios:
        return None
    relatorios.sort(reverse=True)
    caminho = os.path.join(config.RELATORIOS_DIR, relatorios[0])
    with open(caminho, "r", encoding="utf-8") as f:
        return json.load(f)


def gerar_relatorio(planilha: str | None = None, saida: str | None = None) -> str:
    """Gera o relatório final em Markdown."""
    casos = _carregar_casos(planilha)
    defeitos = _carregar_defeitos()
    resultado_pytest = _carregar_resultado_pytest()

    # --- Métricas dos casos de teste (planilha) ---
    total_casos = len(casos)
    status_counter = Counter()
    for caso in casos:
        status = str(caso.get("Status") or "").strip()
        if status:
            status_counter[status] += 1
    aprovados_planilha = (status_counter.get("Aprovado", 0) + status_counter.get("Passou", 0)
                          + status_counter.get("Passed", 0))
    reprovados_planilha = (status_counter.get("Reprovado", 0) + status_counter.get("Falhou", 0)
                           + status_counter.get("Failed", 0))
    pendentes_planilha = total_casos - aprovados_planilha - reprovados_planilha

    funcionalidades_casos = set(caso.get("Funcionalidade") for caso in casos if caso.get("Funcionalidade"))
    funcionalidades_sistema = {"Auth", "Usuários", "Psicólogos", "Sessões", "Mensagens",
                               "Assinatura", "Notificações", "Telefone emergência"}
    funcionalidades_nao_testadas = funcionalidades_sistema - funcionalidades_casos

    # --- Métricas dos defeitos ---
    total_defeitos = len(defeitos)
    defeitos_corrigidos = sum(1 for d in defeitos if d.get("situacao") in ("Corrigido", "Fechado"))
    defeitos_abertos = sum(1 for d in defeitos if d.get("situacao") in ("Aberto", "Em análise", "Reaberto"))
    defeitos_permanentes = sum(1 for d in defeitos
                               if d.get("situacao") == "Fechado" and not d.get("resultado_novo_teste"))
    gravidade_counter = Counter(d.get("gravidade", "Não informado") for d in defeitos)

    # --- Métricas do pytest ---
    pytest_total = pytest_passed = pytest_failed = pytest_skipped = 0
    if resultado_pytest:
        resumo = resultado_pytest.get("summary", {})
        pytest_total = resumo.get("total", 0)
        pytest_passed = resumo.get("passed", 0)
        pytest_failed = resumo.get("failed", 0)
        pytest_skipped = resumo.get("skipped", 0)

    # --- Montagem do relatório ---
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")
    L = []
    L.append("# RELATÓRIO FINAL DE TESTES — CEDRO PLUS")
    L.append("")
    L.append(f"**Data de geração:** {data_geracao}")
    L.append("")
    L.append("---")
    L.append("")
    L.append("## 1. Resumo Executivo")
    L.append("")
    L.append("Este relatório apresenta os resultados dos testes realizados no sistema **Cedro Plus**, ")
    L.append("plataforma de saúde mental que conecta pacientes e psicólogos. Os testes foram ")
    L.append("executados conforme o plano de testes do TCC, cobrindo os módulos de autenticação, ")
    L.append("agendamento de sessões, chat, verificação de CRP, upload de fotos, assinatura e ")
    L.append("gestão de usuários.")
    L.append("")
    L.append("---")
    L.append("")
    L.append("## 2. Métricas de Testes")
    L.append("")
    L.append("### 2.1 Casos de Teste (Planilha)")
    L.append("")
    L.append("| Métrica | Quantidade |")
    L.append("|---|---|")
    L.append(f"| Total de casos de teste | {total_casos} |")
    L.append(f"| Casos aprovados | {aprovados_planilha} |")
    L.append(f"| Casos reprovados | {reprovados_planilha} |")
    L.append(f"| Casos pendentes (não executados) | {pendentes_planilha} |")
    L.append("")
    if total_casos > 0:
        L.append(f"**Taxa de aprovação:** {(aprovados_planilha / total_casos) * 100:.1f}%")
        L.append("")
    L.append("### 2.2 Testes Automatizados (pytest)")
    L.append("")
    L.append("| Métrica | Quantidade |")
    L.append("|---|---|")
    L.append(f"| Total de testes executados | {pytest_total} |")
    L.append(f"| Testes aprovados | {pytest_passed} |")
    L.append(f"| Testes reprovados | {pytest_failed} |")
    L.append(f"| Testes ignorados (skipped) | {pytest_skipped} |")
    L.append("")
    if pytest_total > 0:
        L.append(f"**Taxa de aprovação (pytest):** {(pytest_passed / pytest_total) * 100:.1f}%")
        L.append("")
    L.append("### 2.3 Defeitos Encontrados")
    L.append("")
    L.append("| Métrica | Quantidade |")
    L.append("|---|---|")
    L.append(f"| Total de defeitos encontrados | {total_defeitos} |")
    L.append(f"| Defeitos corrigidos | {defeitos_corrigidos} |")
    L.append(f"| Defeitos abertos (pendentes) | {defeitos_abertos} |")
    L.append(f"| Defeitos permanentes | {defeitos_permanentes} |")
    L.append("")
    L.append("**Distribuição por gravidade:**")
    L.append("")
    L.append("| Gravidade | Quantidade |")
    L.append("|---|---|")
    for g in ["Crítico", "Alto", "Médio", "Baixo"]:
        L.append(f"| {g} | {gravidade_counter.get(g, 0)} |")
    L.append("")
    L.append("### 2.4 Funcionalidades")
    L.append("")
    L.append(f"**Funcionalidades cobertas por casos de teste:** {len(funcionalidades_casos)}")
    L.append("")
    for func in sorted(funcionalidades_casos):
        L.append(f"- {func}")
    L.append("")
    if funcionalidades_nao_testadas:
        L.append(f"**Funcionalidades NÃO testadas:** {len(funcionalidades_nao_testadas)}")
        L.append("")
        for func in sorted(funcionalidades_nao_testadas):
            L.append(f"- {func}")
        L.append("")
    else:
        L.append("**Todas as funcionalidades mapeadas foram cobertas por casos de teste.**")
        L.append("")
    L.append("---")
    L.append("")
    L.append("## 3. Detalhamento dos Defeitos")
    L.append("")
    if not defeitos:
        L.append("Nenhum defeito registrado.")
    else:
        for d in defeitos:
            L.append(f"### {d['id']} — {d['titulo']}")
            L.append("")
            L.append(f"- **Funcionalidade:** {d['funcionalidade']}")
            L.append(f"- **Gravidade:** {d['gravidade']}")
            L.append(f"- **Situação:** {d['situacao']}")
            L.append(f"- **Código do erro:** {d.get('codigo_erro', 'N/A')}")
            L.append(f"- **Descrição:** {d['descricao']}")
            L.append(f"- **Passos para reproduzir:** {d['passos_reproducao']}")
            L.append(f"- **Resultado esperado:** {d['resultado_esperado']}")
            L.append(f"- **Resultado obtido:** {d['resultado_obtido']}")
            L.append(f"- **Responsável:** {d.get('responsavel', 'N/A')}")
            if d.get("data_correcao"):
                L.append(f"- **Data da correção:** {d['data_correcao']}")
            if d.get("resultado_novo_teste"):
                L.append(f"- **Resultado do novo teste:** {d['resultado_novo_teste']}")
            L.append("")
    L.append("---")
    L.append("")
    L.append("## 4. Conclusão Final")
    L.append("")

    # Conclusão baseada nas métricas
    if total_casos == 0 and pytest_total == 0:
        conclusao = ("**Nenhum teste foi executado.** É necessário executar os casos de teste "
                     "da planilha e/ou a suíte automatizada de API para avaliar a qualidade do sistema.")
    else:
        total_executados = max(total_casos, pytest_total)
        total_aprovados = max(aprovados_planilha, pytest_passed)
        if total_executados > 0:
            taxa = (total_aprovados / total_executados) * 100
            if taxa >= 90:
                conclusao = (f"O sistema **Cedro Plus** apresentou uma taxa de aprovação de **{taxa:.1f}%** "
                             f"nos testes executados. A qualidade geral do sistema é considerada **ALTA**, "
                             f"com a maioria dos cenários de teste passando conforme o esperado. "
                             f"Os defeitos encontrados ({total_defeitos}) devem ser corrigidos e re-testados "
                             f"para garantir a estabilidade do sistema.")
            elif taxa >= 70:
                conclusao = (f"O sistema **Cedro Plus** apresentou uma taxa de aprovação de **{taxa:.1f}%** "
                             f"nos testes executados. A qualidade geral é considerada **MODERADA**. "
                             f"Recomenda-se priorizar a correção dos defeitos de gravidade Crítica e Alta "
                             f"antes da entrega final.")
            else:
                conclusao = (f"O sistema **Cedro Plus** apresentou uma taxa de aprovação de **{taxa:.1f}%** "
                             f"nos testes executados. A qualidade geral é considerada **BAIXA**. "
                             f"É necessário corrigir os defeitos encontrados e re-executar os testes "
                             f"antes de considerar o sistema pronto para produção.")
        else:
            conclusao = ("Os testes foram executados, mas não foi possível calcular a taxa de aprovação "
                         "devido à ausência de dados suficientes.")
    L.append(conclusao)
    L.append("")
    L.append("---")
    L.append("")
    L.append("*Relatório gerado automaticamente pelo pacote de automação de testes do TCC.*")

    # Salva o relatório
    if saida is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        saida = os.path.join(config.RELATORIOS_DIR, f"relatorio_final_{timestamp}.md")
    os.makedirs(os.path.dirname(saida), exist_ok=True)
    with open(saida, "w", encoding="utf-8") as f:
        f.write("\n".join(L))
    print(f"[OK] Relatório gerado: {saida}")
    return saida


def main() -> None:
    """Ponto de entrada da CLI."""
    parser = argparse.ArgumentParser(description="Gerador de Relatório Final do Cedro Plus")
    parser.add_argument("--planilha", help="Caminho da planilha de casos de teste", default=None)
    parser.add_argument("--saida", help="Caminho do arquivo de saída", default=None)
    args = parser.parse_args()
    gerar_relatorio(planilha=args.planilha, saida=args.saida)


if __name__ == "__main__":
    main()