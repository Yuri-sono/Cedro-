"""
generate_test_cases.py — Gerador de Casos de Teste para o TCC "Cedro Plus".

Lê as funcionalidades mapeadas no DOCUMENTO_TECNICO_CEDRO.md e gera uma
planilha (.xlsx) com todos os casos de teste, preenchendo os campos
obrigatórios do modelo do PDF do trabalho.

Para CADA funcionalidade principal são gerados automaticamente:
  - Cenário Positivo (dados válidos)
  - Cenário Negativo (dados inválidos)
  - Campo obrigatório vazio
  - Valor no limite (mín/máx)
  - Registro duplicado
  - Acesso não autorizado

As mensagens de erro e regras de negócio foram extraídas diretamente do
código-fonte (ver DOCUMENTO_TECNICO_CEDRO.md, ETAPA 9 — VALIDAÇÕES).
"""

import os
import uuid
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config

# ---------------------------------------------------------------------------
# Constantes extraídas do documento técnico (ETAPA 9 — VALIDAÇÕES)
# ---------------------------------------------------------------------------

# Mensagens de erro exatas do backend (não inventadas)
MSG_LOGIN_INCORRETO = "Email ou senha incorretos"
MSG_EMAIL_DUPLICADO = "Esse email ja ta em uso"
MSG_SENHA_CURTA = "Senha muito curta (min. 6 caracteres)"
MSG_SENHA_SEM_NUMERO = "Precisa ter pelo menos 1 numero"
MSG_SENHA_SEM_ESPECIAL = "Precisa ter pelo menos 1 caractere especial"
MSG_CRP_OBRIGATORIO = "CRP obrigatorio para psicologo"
MSG_CRP_INVALIDO = "CRP invalido. Use o formato 06/123456"
MSG_CRP_DUPLICADO = "Este CRP ja esta cadastrado"
MSG_ESPECIALIDADE_OBRIGATORIA = "Especialidade obrigatoria para psicologo"
MSG_TIPO_PSICOLOGO_OBRIGATORIO = "Tipo de psicologo obrigatorio para psicologo"
MSG_PRECO_OBRIGATORIO = "Valor da consulta obrigatorio para psicologo"
MSG_PSICOLOGO_NAO_ENCONTRADO = "Psicólogo não encontrado"
MSG_LIMITE_SESSOES = "Voce atingiu o limite de 4 sessoes agendadas neste mes no plano gratuito."
MSG_HORARIO_INDISPONIVEL = "Horario indisponivel"
MSG_ACESSO_NEGADO_SESSAO = "Acesso negado. Paciente não corresponde à sessão."
MSG_MENSAGEM_LONGA = "Mensagem muito longa (máx. 2000 caracteres)"
MSG_ARQUIVO_OBRIGATORIO = "Arquivo de imagem obrigatorio"
MSG_IMAGEM_GRANDE = "Imagem muito grande (max. 2MB)"
MSG_FORMATO_INVALIDO = "Formato invalido. Use JPG, PNG ou WebP"
MSG_CRP_FORMATO_VERIFICAR = "Formato de CRP inválido. Use: XX/XXXXXX"
MSG_CRP_CADASTRADO_VERIFICAR = "Este CRP já está cadastrado na plataforma."
MSG_WEBHOOK_NAO_AUTORIZADO = "Webhook nao autorizado"
MSG_PAYLOAD_INVALIDO = "Payload RevenueCat invalido"
MSG_ACESSO_NEGADO = "Acesso negado"
MSG_SENHA_ATUAL_ERRADA = "Senha atual ta errada"
MSG_TOKEN_INVALIDO = "Token invalido ou expirado"
MSG_RECUPERAR_SENHA = "Se o e-mail existir, enviaremos instrucoes"
MSG_CONTA_CRIADA = "Conta criada!"
MSG_SENHA_ALTERADA = "Senha alterada"
MSG_PERFIL_ATUALIZADO = "Perfil atualizado"
MSG_SESSAO_CONFIRMADA = "Sessão confirmada com sucesso"
MSG_DELETADA = "Deletada"
MSG_OK = "ok"

# Credenciais demo (ETAPA 17)
PSICOLOGO_DEMO = "psicologo.demo@cedro.app"
PACIENTE_DEMO = "paciente.demo@cedro.app"
ADMIN_DEMO = "admin@cedro.com"
SENHA_DEMO = "Cedro@123"

# Regras de negócio
CRP_VALIDO = "06/123456"          # CRP do psicólogo demo (seeder Java)
CRP_FORMATO_REGEX = r"\d{2}/\d{5,6}"
SENHA_MIN = 6
MENSAGEM_MAX = 2000
FOTO_MAX_BYTES = 2_000_000        # 2MB (backend)
LIMITE_SESSOES_GRATUITAS = 4
HORARIOS_PADRAO = ["08:00", "09:00", "10:00", "11:00", "14:00", "15:00", "16:00", "17:00", "18:00"]

# ---------------------------------------------------------------------------
# Definição das funcionalidades e seus casos de teste
# ---------------------------------------------------------------------------

def _data_futura(dias: int = 7) -> str:
    """Retorna uma data futura no formato ISO (YYYY-MM-DD)."""
    return (date.today() + timedelta(days=dias)).isoformat()


def _email_unico(prefixo: str) -> str:
    """Gera um email único para evitar conflitos de duplicidade em execuções repetidas."""
    return f"{prefixo}.{uuid.uuid4().hex[:8]}@teste.cedro"


def _gerar_casos_cadastro_paciente() -> list:
    """Casos de teste para a funcionalidade 'Cadastro de Paciente'."""
    email_ok = _email_unico("paciente")
    return [
        {
            "funcionalidade": "Cadastro de Paciente",
            "objetivo": "Verificar o cadastro de um novo paciente com dados válidos",
            "pre_condicao": "Sistema no ar. Email não cadastrado anteriormente.",
            "dados_entrada": f"nome='Maria Teste', email='{email_ok}', senha='Teste@123', "
                             "dataNascimento='1990-01-01', genero='Feminino', telefone='(11) 99999-0000', "
                             "tipoUsuario='paciente'",
            "procedimento": "1. Enviar POST /api/auth/register com os dados acima. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 201 com body {\"message\": \"Conta criada!\"}",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Cadastro de Paciente",
            "objetivo": "Verificar rejeição de cadastro com email duplicado",
            "pre_condicao": "Usuário com o email informado já existe no sistema.",
            "dados_entrada": f"nome='Maria Teste', email='{PACIENTE_DEMO}', senha='Teste@123', "
                             "tipoUsuario='paciente'",
            "procedimento": "1. Enviar POST /api/auth/register com o email do paciente demo. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_EMAIL_DUPLICADO}\"}}",
            "observacoes": "Cenário Negativo — Registro duplicado",
        },
        {
            "funcionalidade": "Cadastro de Paciente",
            "objetivo": "Verificar rejeição de cadastro com campo obrigatório vazio (nome)",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='', email='{_email_unico('paciente')}', senha='Teste@123', "
                             "tipoUsuario='paciente'",
            "procedimento": "1. Enviar POST /api/auth/register com nome vazio. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 400 (validação @NotBlank do campo nome)",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Cadastro de Paciente",
            "objetivo": "Verificar rejeição de cadastro com senha no limite mínimo (6 caracteres)",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Maria Teste', email='{_email_unico('paciente')}', senha='Abc@12', "
                             "tipoUsuario='paciente'",
            "procedimento": "1. Enviar POST /api/auth/register com senha de exatamente 6 caracteres "
                            "contendo número e caractere especial. 2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 201 com body {\"message\": \"Conta criada!\"} — senha de 6 "
                                  "caracteres com número e especial é válida (limite mínimo)",
            "observacoes": "Valor no limite (mín)",
        },
        {
            "funcionalidade": "Cadastro de Paciente",
            "objetivo": "Verificar rejeição de cadastro com senha sem número",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Maria Teste', email='{_email_unico('paciente')}', senha='Teste@abc', "
                             "tipoUsuario='paciente'",
            "procedimento": "1. Enviar POST /api/auth/register com senha sem número. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_SENHA_SEM_NUMERO}\"}}",
            "observacoes": "Cenário Negativo — validação de senha",
        },
        {
            "funcionalidade": "Cadastro de Paciente",
            "objetivo": "Verificar rejeição de cadastro com senha sem caractere especial",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Maria Teste', email='{_email_unico('paciente')}', senha='Teste123', "
                             "tipoUsuario='paciente'",
            "procedimento": "1. Enviar POST /api/auth/register com senha sem caractere especial. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_SENHA_SEM_ESPECIAL}\"}}",
            "observacoes": "Cenário Negativo — validação de senha",
        },
        {
            "funcionalidade": "Cadastro de Paciente",
            "objetivo": "Verificar que a API pública não permite criar usuário admin",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Admin Fake', email='{_email_unico('adminfake')}', senha='Teste@123', "
                             "tipoUsuario='admin'",
            "procedimento": "1. Enviar POST /api/auth/register com tipoUsuario='admin'. "
                            "2. Verificar o código HTTP e o tipo de usuário criado.",
            "resultado_esperado": "HTTP 201 — o backend força tipoUsuario='paciente' "
                                  "(admin não pode ser criado via API pública)",
            "observacoes": "Acesso não autorizado — tentativa de escalonamento de privilégio",
        },
    ]


def _gerar_casos_cadastro_psicologo() -> list:
    """Casos de teste para a funcionalidade 'Cadastro de Psicólogo'."""
    email_ok = _email_unico("psicologo")
    crp_ok = f"06/{uuid.uuid4().hex[:6].upper()}"
    return [
        {
            "funcionalidade": "Cadastro de Psicólogo",
            "objetivo": "Verificar o cadastro de um novo psicólogo com dados válidos",
            "pre_condicao": "Sistema no ar. CRP não cadastrado anteriormente.",
            "dados_entrada": f"nome='Dr. João Teste', email='{email_ok}', senha='Teste@123', "
                             f"tipoUsuario='psicologo', crp='{crp_ok}', especialidade='Psicologia Clínica', "
                             "tipoPsicologo='Terapia Individual', precoSessao=150.00",
            "procedimento": "1. Enviar POST /api/auth/register com os dados acima. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 201 com body {\"message\": \"Conta criada!\"}",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Cadastro de Psicólogo",
            "objetivo": "Verificar rejeição de cadastro de psicólogo sem CRP",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Dr. João Teste', email='{_email_unico('psicologo')}', senha='Teste@123', "
                             "tipoUsuario='psicologo', especialidade='Psicologia Clínica', "
                             "tipoPsicologo='Terapia Individual', precoSessao=150.00",
            "procedimento": "1. Enviar POST /api/auth/register sem o campo crp. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_CRP_OBRIGATORIO}\"}}",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Cadastro de Psicólogo",
            "objetivo": "Verificar rejeição de cadastro com CRP em formato inválido",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Dr. João Teste', email='{_email_unico('psicologo')}', senha='Teste@123', "
                             "tipoUsuario='psicologo', crp='12345', especialidade='Psicologia Clínica', "
                             "tipoPsicologo='Terapia Individual', precoSessao=150.00",
            "procedimento": "1. Enviar POST /api/auth/register com crp='12345' (formato inválido). "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_CRP_INVALIDO}\"}}",
            "observacoes": "Cenário Negativo — formato inválido",
        },
        {
            "funcionalidade": "Cadastro de Psicólogo",
            "objetivo": "Verificar rejeição de cadastro com CRP duplicado",
            "pre_condicao": "CRP 06/123456 já cadastrado (psicólogo demo do seeder).",
            "dados_entrada": f"nome='Dr. João Teste', email='{_email_unico('psicologo')}', senha='Teste@123', "
                             "tipoUsuario='psicologo', crp='06/123456', especialidade='Psicologia Clínica', "
                             "tipoPsicologo='Terapia Individual', precoSessao=150.00",
            "procedimento": "1. Enviar POST /api/auth/register com crp='06/123456' (já cadastrado). "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_CRP_DUPLICADO}\"}}",
            "observacoes": "Registro duplicado",
        },
        {
            "funcionalidade": "Cadastro de Psicólogo",
            "objetivo": "Verificar rejeição de cadastro sem especialidade",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Dr. João Teste', email='{_email_unico('psicologo')}', senha='Teste@123', "
                             f"tipoUsuario='psicologo', crp='{crp_ok}', tipoPsicologo='Terapia Individual', "
                             "precoSessao=150.00",
            "procedimento": "1. Enviar POST /api/auth/register sem o campo especialidade. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_ESPECIALIDADE_OBRIGATORIA}\"}}",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Cadastro de Psicólogo",
            "objetivo": "Verificar rejeição de cadastro sem tipo de psicólogo",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Dr. João Teste', email='{_email_unico('psicologo')}', senha='Teste@123', "
                             f"tipoUsuario='psicologo', crp='{crp_ok}', especialidade='Psicologia Clínica', "
                             "precoSessao=150.00",
            "procedimento": "1. Enviar POST /api/auth/register sem o campo tipoPsicologo. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_TIPO_PSICOLOGO_OBRIGATORIO}\"}}",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Cadastro de Psicólogo",
            "objetivo": "Verificar rejeição de cadastro sem preço da sessão",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Dr. João Teste', email='{_email_unico('psicologo')}', senha='Teste@123', "
                             f"tipoUsuario='psicologo', crp='{crp_ok}', especialidade='Psicologia Clínica', "
                             "tipoPsicologo='Terapia Individual'",
            "procedimento": "1. Enviar POST /api/auth/register sem o campo precoSessao. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_PRECO_OBRIGATORIO}\"}}",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Cadastro de Psicólogo",
            "objetivo": "Verificar rejeição de cadastro com preço zero ou negativo",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"nome='Dr. João Teste', email='{_email_unico('psicologo')}', senha='Teste@123', "
                             f"tipoUsuario='psicologo', crp='{crp_ok}', especialidade='Psicologia Clínica', "
                             "tipoPsicologo='Terapia Individual', precoSessao=0",
            "procedimento": "1. Enviar POST /api/auth/register com precoSessao=0. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_PRECO_OBRIGATORIO}\"}}",
            "observacoes": "Valor no limite (mín) — preço deve ser > 0",
        },
    ]


def _gerar_casos_login() -> list:
    """Casos de teste para a funcionalidade 'Login'."""
    return [
        {
            "funcionalidade": "Login",
            "objetivo": "Verificar login com credenciais válidas",
            "pre_condicao": "Usuário demo ativo no sistema (seeder Java).",
            "dados_entrada": f"email='{PACIENTE_DEMO}', senha='{SENHA_DEMO}'",
            "procedimento": "1. Enviar POST /api/auth/login com as credenciais do paciente demo. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com body contendo 'token' (JWT) e 'usuario' com "
                                  "tipoUsuario='paciente'",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Login",
            "objetivo": "Verificar rejeição de login com senha incorreta",
            "pre_condicao": "Usuário demo ativo no sistema.",
            "dados_entrada": f"email='{PACIENTE_DEMO}', senha='SenhaErrada@1'",
            "procedimento": "1. Enviar POST /api/auth/login com senha incorreta. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_LOGIN_INCORRETO}\"}}",
            "observacoes": "Cenário Negativo",
        },
        {
            "funcionalidade": "Login",
            "objetivo": "Verificar rejeição de login com email inexistente",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"email='{_email_unico('naoexiste')}', senha='{SENHA_DEMO}'",
            "procedimento": "1. Enviar POST /api/auth/login com email não cadastrado. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_LOGIN_INCORRETO}\"}} "
                                  "(mensagem genérica — não revela se email existe)",
            "observacoes": "Cenário Negativo",
        },
        {
            "funcionalidade": "Login",
            "objetivo": "Verificar rejeição de login com usuário inativo",
            "pre_condicao": "Usuário com ativo=false no banco (desativado por admin).",
            "dados_entrada": "email='<usuario_inativo>', senha='<senha_correta>'",
            "procedimento": "1. Enviar POST /api/auth/login com credenciais de usuário inativo. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_LOGIN_INCORRETO}\"}} "
                                  "(findByEmailIgnoreCaseAndAtivoTrue não encontra o usuário)",
            "observacoes": "Cenário Negativo — usuário inativo",
        },
        {
            "funcionalidade": "Login",
            "objetivo": "Verificar rejeição de login com campo obrigatório vazio (email)",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "email='', senha='Cedro@123'",
            "procedimento": "1. Enviar POST /api/auth/login com email vazio. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 400 (validação @NotBlank + @Email do campo email)",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Login",
            "objetivo": "Verificar rejeição de login com campo obrigatório vazio (senha)",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"email='{PACIENTE_DEMO}', senha=''",
            "procedimento": "1. Enviar POST /api/auth/login com senha vazia. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 400 (validação @NotBlank do campo senha)",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Login",
            "objetivo": "Verificar login com email em caixa alta (normalização)",
            "pre_condicao": "Usuário demo ativo no sistema.",
            "dados_entrada": f"email='{PACIENTE_DEMO.upper()}', senha='{SENHA_DEMO}'",
            "procedimento": "1. Enviar POST /api/auth/login com email em MAIÚSCULAS. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 200 — o backend normaliza o email (trim + lowercase) "
                                  "antes de buscar no banco",
            "observacoes": "Valor no limite — normalização de email",
        },
        {
            "funcionalidade": "Login",
            "objetivo": "Verificar que psicólogo não consegue acessar área de paciente via login comum",
            "pre_condicao": "Psicólogo demo ativo no sistema.",
            "dados_entrada": f"email='{PSICOLOGO_DEMO}', senha='{SENHA_DEMO}'",
            "procedimento": "1. Enviar POST /api/auth/login com credenciais do psicólogo demo. "
                            "2. Verificar o tipoUsuario retornado.",
            "resultado_esperado": "HTTP 200 com tipoUsuario='psicologo' — o frontend "
                                  "(LoginPsicologo.jsx) bloqueia com 'Esta conta não é de psicólogo. "
                                  "Use o login de paciente.' se tentar usar como paciente",
            "observacoes": "Acesso não autorizado — controle de perfil no frontend",
        },
    ]


def _gerar_casos_agendamento() -> list:
    """Casos de teste para a funcionalidade 'Agendamento de Sessão'."""
    data_futura = _data_futura(7)
    return [
        {
            "funcionalidade": "Agendamento de Sessão",
            "objetivo": "Verificar agendamento de sessão com dados válidos",
            "pre_condicao": "Paciente autenticado (JWT). Psicólogo válido com horário livre.",
            "dados_entrada": f"psicologoId=<id_psicologo_demo>, dataSessao='{data_futura}T10:00:00', "
                             "duracao=60, observacoes='Sessão de teste'",
            "procedimento": "1. Obter token JWT do paciente demo (POST /api/auth/login). "
                            "2. Enviar POST /api/sessoes com header Authorization: Bearer <token>. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 201 com a Sessao criada (pacienteId = id do token, "
                                  "valor = precoSessao do psicólogo)",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Agendamento de Sessão",
            "objetivo": "Verificar rejeição de agendamento com psicólogo inexistente",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": f"psicologoId=999999, dataSessao='{data_futura}T10:00:00', duracao=60",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/sessoes com psicologoId inexistente. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_PSICOLOGO_NAO_ENCONTRADO}\"}}",
            "observacoes": "Cenário Negativo",
        },
        {
            "funcionalidade": "Agendamento de Sessão",
            "objetivo": "Verificar rejeição de agendamento com horário ocupado",
            "pre_condicao": "Paciente autenticado. Já existe sessão não-cancelada do psicólogo "
                            "no mesmo horário.",
            "dados_entrada": f"psicologoId=<id_psicologo_demo>, dataSessao='{data_futura}T10:00:00', "
                             "duracao=60 (mesmo horário de sessão existente)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/sessoes com horário já ocupado. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_HORARIO_INDISPONIVEL}\"}}",
            "observacoes": "Registro duplicado — horário do psicólogo",
        },
        {
            "funcionalidade": "Agendamento de Sessão",
            "objetivo": "Verificar rejeição de agendamento sem psicologoId (campo obrigatório)",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": f"dataSessao='{data_futura}T10:00:00', duracao=60",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/sessoes sem o campo psicologoId. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 400 (validação @NotNull — 'psicologoId é obrigatório')",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Agendamento de Sessão",
            "objetivo": "Verificar rejeição de agendamento sem dataSessao (campo obrigatório)",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "psicologoId=<id_psicologo_demo>, duracao=60",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/sessoes sem o campo dataSessao. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 400 (validação @NotNull — 'dataSessao é obrigatória')",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Agendamento de Sessão",
            "objetivo": "Verificar rejeição de agendamento sem token JWT (acesso não autorizado)",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"psicologoId=<id_psicologo_demo>, dataSessao='{data_futura}T10:00:00', "
                             "duracao=60 (sem header Authorization)",
            "procedimento": "1. Enviar POST /api/sessoes sem header Authorization. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 401 (endpoint exige JWT — SecurityConfig)",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Agendamento de Sessão",
            "objetivo": "Verificar que o backend ignora o pacienteId enviado no body",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": f"pacienteId=999999, psicologoId=<id_psicologo_demo>, "
                             f"dataSessao='{data_futura}T11:00:00', duracao=60",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/sessoes com pacienteId diferente do token. "
                            "3. Verificar o pacienteId da sessão criada.",
            "resultado_esperado": "HTTP 201 — o controller força request.setPacienteId(userId) "
                                  "do token (ignora pacienteId do body)",
            "observacoes": "Regra de negócio — segurança",
        },
        {
            "funcionalidade": "Agendamento de Sessão",
            "objetivo": "Verificar o limite de 4 sessões gratuitas por mês",
            "pre_condicao": "Paciente não-premium com 4 sessões não-canceladas criadas no mês corrente.",
            "dados_entrada": f"psicologoId=<id_psicologo_demo>, dataSessao='{data_futura}T14:00:00', "
                             "duracao=60 (5ª sessão do mês)",
            "procedimento": "1. Obter token JWT do paciente demo (não-premium). "
                            "2. Enviar POST /api/sessoes pela 5ª vez no mês. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 403 com body {{\"error\": \"{MSG_LIMITE_SESSOES}\"}}",
            "observacoes": "Valor no limite (máx) — 4 sessões/mês para plano gratuito",
        },
    ]


def _gerar_casos_pagamento() -> list:
    """Casos de teste para a funcionalidade 'Pagamento e Confirmação de Sessão'."""
    return [
        {
            "funcionalidade": "Pagamento e Confirmação de Sessão",
            "objetivo": "Verificar confirmação de pagamento pelo paciente dono da sessão",
            "pre_condicao": "Paciente autenticado (JWT). Sessão criada com status 'pendente' "
                            "pertencente ao paciente.",
            "dados_entrada": "sessaoId=<id_sessao_do_paciente>",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/sessoes/{id}/confirmar-pagamento com o token. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 200 com body contendo \"message\": \"{MSG_SESSAO_CONFIRMADA}\" "
                                  "e a sessão com status 'agendada'",
            "observacoes": "Cenário Positivo — PIX fake do frontend não é testado; "
                           "apenas o endpoint real de confirmação",
        },
        {
            "funcionalidade": "Pagamento e Confirmação de Sessão",
            "objetivo": "Verificar rejeição de confirmação de pagamento por outro paciente",
            "pre_condicao": "Paciente A autenticado. Sessão pertencente ao Paciente B.",
            "dados_entrada": "sessaoId=<id_sessao_do_paciente_B> (token do paciente A)",
            "procedimento": "1. Obter token JWT do paciente A. "
                            "2. Enviar POST /api/sessoes/{id}/confirmar-pagamento da sessão do paciente B. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 403 com body {{\"error\": \"{MSG_ACESSO_NEGADO_SESSAO}\"}}",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Pagamento e Confirmação de Sessão",
            "objetivo": "Verificar rejeição de confirmação de pagamento sem token JWT",
            "pre_condicao": "Sessão existente no sistema.",
            "dados_entrada": "sessaoId=<id_sessao> (sem header Authorization)",
            "procedimento": "1. Enviar POST /api/sessoes/{id}/confirmar-pagamento sem token. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 401 (endpoint exige JWT)",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Pagamento e Confirmação de Sessão",
            "objetivo": "Verificar rejeição de confirmação de pagamento de sessão inexistente",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "sessaoId=999999",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/sessoes/999999/confirmar-pagamento. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 400 (sessão não encontrada — RuntimeException "
                                  "'Não encontrada' tratada pelo GlobalExceptionHandler)",
            "observacoes": "Cenário Negativo",
        },
        {
            "funcionalidade": "Pagamento e Confirmação de Sessão",
            "objetivo": "Verificar o link de reunião antes da janela de liberação (15 min)",
            "pre_condicao": "Paciente autenticado. Sessão confirmada com data futura "
                            "(mais de 15 min de antecedência).",
            "dados_entrada": "sessaoId=<id_sessao_confirmada>",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar GET /api/sessoes/{id}/link-reuniao. "
                            "3. Verificar o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com body {\"liberado\": false, "
                                  "\"disponivelEm\": \"<ISO datetime>\"}",
            "observacoes": "Valor no limite — janela de liberação de 15 minutos "
                           "(GOOGLE_MEET_RELEASE_MINUTES_BEFORE)",
        },
        {
            "funcionalidade": "Pagamento e Confirmação de Sessão",
            "objetivo": "Verificar acesso ao link de reunião de outro usuário",
            "pre_condicao": "Paciente A autenticado. Sessão pertencente ao Paciente B.",
            "dados_entrada": "sessaoId=<id_sessao_do_paciente_B> (token do paciente A)",
            "procedimento": "1. Obter token JWT do paciente A. "
                            "2. Enviar GET /api/sessoes/{id}/link-reuniao da sessão do paciente B. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 403 (apenas admin, paciente ou psicólogo da sessão)",
            "observacoes": "Acesso não autorizado",
        },
    ]


def _gerar_casos_chat() -> list:
    """Casos de teste para a funcionalidade 'Chat (Mensagens)'."""
    return [
        {
            "funcionalidade": "Chat (Mensagens)",
            "objetivo": "Verificar envio de mensagem com dados válidos",
            "pre_condicao": "Paciente e psicólogo autenticados (JWTs).",
            "dados_entrada": "destinatarioId=<id_psicologo_demo>, mensagem='Olá, gostaria de agendar uma sessão'",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/mensagens com header Authorization: Bearer <token>. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com a Mensagem criada (remetenteId = id do token)",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Chat (Mensagens)",
            "objetivo": "Verificar rejeição de mensagem com mais de 2000 caracteres",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "destinatarioId=<id_psicologo_demo>, mensagem='<2001 caracteres>'",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/mensagens com mensagem de 2001 caracteres. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_MENSAGEM_LONGA}\"}}",
            "observacoes": "Valor no limite (máx) — 2000 caracteres",
        },
        {
            "funcionalidade": "Chat (Mensagens)",
            "objetivo": "Verificar envio de mensagem com exatamente 2000 caracteres (limite)",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "destinatarioId=<id_psicologo_demo>, mensagem='<2000 caracteres>'",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/mensagens com mensagem de exatamente 2000 caracteres. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 200 — mensagem aceita no limite máximo",
            "observacoes": "Valor no limite (máx)",
        },
        {
            "funcionalidade": "Chat (Mensagens)",
            "objetivo": "Verificar rejeição de mensagem sem destinatarioId (campo obrigatório)",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "mensagem='Olá' (sem destinatarioId)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/mensagens sem o campo destinatarioId. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 400 (validação @NotNull do campo destinatarioId)",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Chat (Mensagens)",
            "objetivo": "Verificar rejeição de mensagem sem token JWT",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "destinatarioId=<id_psicologo_demo>, mensagem='Olá' (sem token)",
            "procedimento": "1. Enviar POST /api/mensagens sem header Authorization. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 401 (endpoint exige JWT)",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Chat (Mensagens)",
            "objetivo": "Verificar que apenas o destinatário pode marcar mensagem como lida",
            "pre_condicao": "Mensagem existente entre paciente e psicólogo.",
            "dados_entrada": "mensagemId=<id_mensagem> (token do remetente, não do destinatário)",
            "procedimento": "1. Obter token JWT do remetente da mensagem. "
                            "2. Enviar PUT /api/mensagens/{id}/lida com o token do remetente. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 403 (apenas o destinatário pode marcar como lida)",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Chat (Mensagens)",
            "objetivo": "Verificar listagem de conversas do usuário autenticado",
            "pre_condicao": "Paciente autenticado com mensagens trocadas.",
            "dados_entrada": "Nenhum (GET)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar GET /api/mensagens/conversas com o token. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com lista de ConversaResumo "
                                  "{userId, nome, fotoUrl, ultimaMensagem, dataUltimaMensagem, "
                                  "naoLidas, mensagemEnviada}",
            "observacoes": "Cenário Positivo",
        },
    ]


def _gerar_casos_verificar_crp() -> list:
    """Casos de teste para a funcionalidade 'Verificação de CRP'."""
    crp_disponivel = f"06/{uuid.uuid4().hex[:6].upper()}"
    return [
        {
            "funcionalidade": "Verificação de CRP",
            "objetivo": "Verificar CRP disponível para cadastro",
            "pre_condicao": "Sistema no ar. CRP não cadastrado.",
            "dados_entrada": f"crp='{crp_disponivel}'",
            "procedimento": "1. Enviar GET /api/psicologos/verificar-crp?crp=<crp>. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com body {\"valido\": true, "
                                  "\"mensagem\": \"CRP disponível para cadastro\"}",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Verificação de CRP",
            "objetivo": "Verificar rejeição de CRP em formato inválido",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "crp='12345'",
            "procedimento": "1. Enviar GET /api/psicologos/verificar-crp?crp=12345. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"valido\": false, "
                                  f"\"mensagem\": \"{MSG_CRP_FORMATO_VERIFICAR}\"}}",
            "observacoes": "Cenário Negativo — formato inválido",
        },
        {
            "funcionalidade": "Verificação de CRP",
            "objetivo": "Verificar rejeição de CRP já cadastrado",
            "pre_condicao": "CRP 06/123456 já cadastrado (psicólogo demo do seeder).",
            "dados_entrada": "crp='06/123456'",
            "procedimento": "1. Enviar GET /api/psicologos/verificar-crp?crp=06/123456. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 409 com body {{\"valido\": false, "
                                  f"\"mensagem\": \"{MSG_CRP_CADASTRADO_VERIFICAR}\"}}",
            "observacoes": "Registro duplicado",
        },
        {
            "funcionalidade": "Verificação de CRP",
            "objetivo": "Verificar CRP com formato no limite (5 dígitos após a barra)",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"crp='06/{uuid.uuid4().hex[:5].upper()}'",
            "procedimento": "1. Enviar GET /api/psicologos/verificar-crp com CRP de 5 dígitos. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 200 (formato \\d{2}/\\d{5,6} aceita 5 ou 6 dígitos) "
                                  "ou HTTP 409 se já cadastrado",
            "observacoes": "Valor no limite (mín) — 5 dígitos",
        },
        {
            "funcionalidade": "Verificação de CRP",
            "objetivo": "Verificar CRP com formato no limite (6 dígitos após a barra)",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"crp='06/{uuid.uuid4().hex[:6].upper()}'",
            "procedimento": "1. Enviar GET /api/psicologos/verificar-crp com CRP de 6 dígitos. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 200 (formato \\d{2}/\\d{5,6} aceita 5 ou 6 dígitos) "
                                  "ou HTTP 409 se já cadastrado",
            "observacoes": "Valor no limite (máx) — 6 dígitos",
        },
        {
            "funcionalidade": "Verificação de CRP",
            "objetivo": "Verificar CRP com 7 dígitos (acima do limite)",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "crp='06/1234567'",
            "procedimento": "1. Enviar GET /api/psicologos/verificar-crp?crp=06/1234567. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"valido\": false, "
                                  f"\"mensagem\": \"{MSG_CRP_FORMATO_VERIFICAR}\"}}",
            "observacoes": "Valor no limite (máx+1) — 7 dígitos é inválido",
        },
    ]


def _gerar_casos_alterar_senha() -> list:
    """Casos de teste para a funcionalidade 'Alteração de Senha'."""
    return [
        {
            "funcionalidade": "Alteração de Senha",
            "objetivo": "Verificar alteração de senha com dados válidos",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": f"senhaAtual='{SENHA_DEMO}', novaSenha='NovaSenha@123'",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar PUT /api/auth/alterar-senha com o token. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 200 com body {{\"message\": \"{MSG_SENHA_ALTERADA}\"}}",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Alteração de Senha",
            "objetivo": "Verificar rejeição de alteração com senha atual incorreta",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "senhaAtual='SenhaErrada@1', novaSenha='NovaSenha@123'",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar PUT /api/auth/alterar-senha com senha atual errada. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_SENHA_ATUAL_ERRADA}\"}}",
            "observacoes": "Cenário Negativo",
        },
        {
            "funcionalidade": "Alteração de Senha",
            "objetivo": "Verificar rejeição de nova senha sem número",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": f"senhaAtual='{SENHA_DEMO}', novaSenha='NovaSenha@abc'",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar PUT /api/auth/alterar-senha com nova senha sem número. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_SENHA_SEM_NUMERO}\"}}",
            "observacoes": "Cenário Negativo — validação de força da senha",
        },
        {
            "funcionalidade": "Alteração de Senha",
            "objetivo": "Verificar rejeição de nova senha sem caractere especial",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": f"senhaAtual='{SENHA_DEMO}', novaSenha='NovaSenha123'",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar PUT /api/auth/alterar-senha com nova senha sem especial. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_SENHA_SEM_ESPECIAL}\"}}",
            "observacoes": "Cenário Negativo — validação de força da senha",
        },
        {
            "funcionalidade": "Alteração de Senha",
            "objetivo": "Verificar rejeição de nova senha com menos de 6 caracteres",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": f"senhaAtual='{SENHA_DEMO}', novaSenha='Ab@12'",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar PUT /api/auth/alterar-senha com nova senha de 5 caracteres. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_SENHA_CURTA}\"}}",
            "observacoes": "Valor no limite (mín-1) — 5 caracteres",
        },
        {
            "funcionalidade": "Alteração de Senha",
            "objetivo": "Verificar rejeição de alteração de senha sem token JWT",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"senhaAtual='{SENHA_DEMO}', novaSenha='NovaSenha@123' (sem token)",
            "procedimento": "1. Enviar PUT /api/auth/alterar-senha sem header Authorization. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 401 (endpoint exige JWT)",
            "observacoes": "Acesso não autorizado",
        },
    ]


def _gerar_casos_upload_foto() -> list:
    """Casos de teste para a funcionalidade 'Upload de Foto de Perfil'."""
    return [
        {
            "funcionalidade": "Upload de Foto de Perfil",
            "objetivo": "Verificar upload de imagem JPG válida",
            "pre_condicao": "Paciente autenticado (JWT). Arquivo JPG válido de até 2MB.",
            "dados_entrada": "file=<foto.jpg> (multipart/form-data, < 2MB, content-type image/jpeg)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/auth/foto-perfil-upload com o arquivo e o token. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com body {\"message\": \"Foto atualizada\", "
                                  "\"fotoUrl\": \"<url>\"}",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Upload de Foto de Perfil",
            "objetivo": "Verificar rejeição de upload de imagem maior que 2MB",
            "pre_condicao": "Paciente autenticado (JWT). Arquivo JPG de 2.1MB.",
            "dados_entrada": "file=<foto_grande.jpg> (multipart/form-data, > 2MB)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/auth/foto-perfil-upload com arquivo > 2MB. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_IMAGEM_GRANDE}\"}}",
            "observacoes": "Valor no limite (máx) — 2MB (2_000_000 bytes)",
        },
        {
            "funcionalidade": "Upload de Foto de Perfil",
            "objetivo": "Verificar rejeição de upload de arquivo sem imagem",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "file=<arquivo.txt> (content-type text/plain)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/auth/foto-perfil-upload com arquivo .txt. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_FORMATO_INVALIDO}\"}}",
            "observacoes": "Cenário Negativo — formato inválido",
        },
        {
            "funcionalidade": "Upload de Foto de Perfil",
            "objetivo": "Verificar rejeição de upload sem arquivo",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "file=<vazio> (multipart/form-data sem arquivo)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/auth/foto-perfil-upload sem arquivo. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_ARQUIVO_OBRIGATORIO}\"}}",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Upload de Foto de Perfil",
            "objetivo": "Verificar rejeição de upload sem token JWT",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "file=<foto.jpg> (sem header Authorization)",
            "procedimento": "1. Enviar POST /api/auth/foto-perfil-upload sem token. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 401 (endpoint exige JWT)",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Upload de Foto de Perfil",
            "objetivo": "Verificar upload de imagem PNG válida",
            "pre_condicao": "Paciente autenticado (JWT). Arquivo PNG válido de até 2MB.",
            "dados_entrada": "file=<foto.png> (multipart/form-data, < 2MB, content-type image/png)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar POST /api/auth/foto-perfil-upload com arquivo PNG. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 200 com body {\"message\": \"Foto atualizada\", "
                                  "\"fotoUrl\": \"<url>\"}",
            "observacoes": "Cenário Positivo — formatos aceitos: JPG, PNG, WebP",
        },
    ]


def _gerar_casos_assinatura() -> list:
    """Casos de teste para a funcionalidade 'Assinatura Premium (RevenueCat)'."""
    return [
        {
            "funcionalidade": "Assinatura Premium (RevenueCat)",
            "objetivo": "Verificar processamento de webhook autorizado",
            "pre_condicao": "REVENUECAT_WEBHOOK_SECRET configurado no backend.",
            "dados_entrada": "Header Authorization=<secret>. Body: "
                             "{\"event\": {\"type\": \"INITIAL_PURCHASE\", "
                             "\"app_user_id\": \"<id_usuario>\", \"expiration_at_ms\": 1750000000000, "
                             "\"product_id\": \"premium_mensal\", \"transaction_id\": \"tx123\"}}",
            "procedimento": "1. Enviar POST /api/assinatura/webhook com header Authorization correto. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com body {\"message\": \"Webhook processado\"}",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Assinatura Premium (RevenueCat)",
            "objetivo": "Verificar rejeição de webhook não autorizado",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "Header Authorization=<secret_errado>",
            "procedimento": "1. Enviar POST /api/assinatura/webhook com header Authorization incorreto. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 401 com body {{\"error\": \"{MSG_WEBHOOK_NAO_AUTORIZADO}\"}}",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Assinatura Premium (RevenueCat)",
            "objetivo": "Verificar rejeição de webhook com payload inválido",
            "pre_condicao": "REVENUECAT_WEBHOOK_SECRET configurado no backend.",
            "dados_entrada": "Header Authorization=<secret>. Body: {\"event\": {}}",
            "procedimento": "1. Enviar POST /api/assinatura/webhook com payload sem type/app_user_id. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_PAYLOAD_INVALIDO}\"}}",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Assinatura Premium (RevenueCat)",
            "objetivo": "Verificar status de assinatura de usuário free",
            "pre_condicao": "Paciente autenticado sem assinatura ativa.",
            "dados_entrada": "Nenhum (GET)",
            "procedimento": "1. Obter token JWT do paciente demo (não-premium). "
                            "2. Enviar GET /api/assinatura/status com o token. "
                            "3. Verificar o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com body {\"isPremium\": false, "
                                  "\"chamadasRealizadas\": <count>, \"limiteMensal\": 4}",
            "observacoes": "Cenário Positivo — limite gratuito de 4 sessões/mês",
        },
        {
            "funcionalidade": "Assinatura Premium (RevenueCat)",
            "objetivo": "Verificar status de assinatura sem token JWT",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "Nenhum (GET, sem token)",
            "procedimento": "1. Enviar GET /api/assinatura/status sem header Authorization. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 401 (endpoint exige JWT)",
            "observacoes": "Acesso não autorizado",
        },
    ]


def _gerar_casos_usuarios() -> list:
    """Casos de teste para a funcionalidade 'Gestão de Usuários (Admin)'."""
    return [
        {
            "funcionalidade": "Gestão de Usuários (Admin)",
            "objetivo": "Verificar listagem de todos os usuários por admin",
            "pre_condicao": "Admin autenticado (JWT).",
            "dados_entrada": "Nenhum (GET)",
            "procedimento": "1. Obter token JWT do admin (admin@cedro.com). "
                            "2. Enviar GET /api/usuarios com o token. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 200 com lista de todos os usuários",
            "observacoes": "Cenário Positivo — NOTA: endpoint expõe senhaHash (vulnerabilidade conhecida)",
        },
        {
            "funcionalidade": "Gestão de Usuários (Admin)",
            "objetivo": "Verificar rejeição de listagem de usuários por não-admin",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "Nenhum (GET, token de paciente)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar GET /api/usuarios com o token do paciente. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 403 com body {{\"error\": \"{MSG_ACESSO_NEGADO}\"}}",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Gestão de Usuários (Admin)",
            "objetivo": "Verificar que usuário comum só pode ver a si mesmo",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "id=<id_de_outro_usuario> (token do paciente)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar GET /api/usuarios/{id} de outro usuário. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 403 (se não admin e id não corresponde)",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Gestão de Usuários (Admin)",
            "objetivo": "Verificar ativação/desativação de usuário por admin",
            "pre_condicao": "Admin autenticado (JWT).",
            "dados_entrada": "id=<id_usuario>, body={\"ativo\": false}",
            "procedimento": "1. Obter token JWT do admin. "
                            "2. Enviar PUT /api/usuarios/{id}/ativar com ativo=false. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 200 com body {{\"message\": \"{MSG_OK}\"}}",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Gestão de Usuários (Admin)",
            "objetivo": "Verificar rejeição de desativação de usuário por não-admin",
            "pre_condicao": "Paciente autenticado (JWT).",
            "dados_entrada": "id=<id_usuario>, body={\"ativo\": false} (token de paciente)",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Enviar PUT /api/usuarios/{id}/ativar com o token do paciente. "
                            "3. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 403 (apenas admin)",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Gestão de Usuários (Admin)",
            "objetivo": "Verificar exclusão de usuário por admin",
            "pre_condicao": "Admin autenticado (JWT). Usuário de teste criado.",
            "dados_entrada": "id=<id_usuario_teste>",
            "procedimento": "1. Obter token JWT do admin. "
                            "2. Enviar DELETE /api/usuarios/{id} do usuário de teste. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 200 com body {{\"message\": \"{MSG_DELETADA}\"}}",
            "observacoes": "Cenário Positivo",
        },
    ]


def _gerar_casos_psicologos() -> list:
    """Casos de teste para a funcionalidade 'Lista e Detalhe de Psicólogos'."""
    return [
        {
            "funcionalidade": "Lista e Detalhe de Psicólogos",
            "objetivo": "Verificar listagem pública de psicólogos",
            "pre_condicao": "Sistema no ar. Psicólogos cadastrados.",
            "dados_entrada": "Nenhum (GET, público)",
            "procedimento": "1. Enviar GET /api/psicologos sem autenticação. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com lista de DTOs públicos "
                                  "{id, nome, especialidade, tipoPsicologo, bio, precoSessao, "
                                  "avaliacao, fotoUrl} — sem email, telefone ou senha",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Lista e Detalhe de Psicólogos",
            "objetivo": "Verificar detalhe de psicólogo inexistente",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "id=999999",
            "procedimento": "1. Enviar GET /api/psicologos/999999. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 400 (RuntimeException 'Não encontrado' tratada pelo "
                                  "GlobalExceptionHandler)",
            "observacoes": "Cenário Negativo",
        },
        {
            "funcionalidade": "Lista e Detalhe de Psicólogos",
            "objetivo": "Verificar detalhe de usuário que não é psicólogo",
            "pre_condicao": "Sistema no ar. Paciente demo cadastrado.",
            "dados_entrada": "id=<id_paciente_demo>",
            "procedimento": "1. Enviar GET /api/psicologos/{id} do paciente demo. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 400 com body {\"error\": \"Não é psicólogo\"}",
            "observacoes": "Cenário Negativo",
        },
        {
            "funcionalidade": "Lista e Detalhe de Psicólogos",
            "objetivo": "Verificar filtro de psicólogos por área de interesse do paciente",
            "pre_condicao": "Paciente autenticado com areaInteresse definida.",
            "dados_entrada": "Nenhum (GET, token de paciente com areaInteresse)",
            "procedimento": "1. Obter token JWT do paciente com areaInteresse. "
                            "2. Enviar GET /api/psicologos com o token. "
                            "3. Verificar que a lista é filtrada por correspondência de tags.",
            "resultado_esperado": "HTTP 200 — apenas psicólogos cujo tipoPsicologo corresponde "
                                  "à areaInteresse do paciente (normalização de tags)",
            "observacoes": "Cenário Positivo — filtro por interesse",
        },
        {
            "funcionalidade": "Lista e Detalhe de Psicólogos",
            "objetivo": "Verificar financeiro do psicólogo",
            "pre_condicao": "Psicólogo autenticado (JWT).",
            "dados_entrada": "periodo='mes'",
            "procedimento": "1. Obter token JWT do psicólogo demo. "
                            "2. Enviar GET /api/psicologos/financeiro?periodo=mes. "
                            "3. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 200 com faturamentoMes, consultasRealizadas, ticketMedio "
                                  "e transacoes (últimas 20 não-canceladas)",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Lista e Detalhe de Psicólogos",
            "objetivo": "Verificar rejeição de acesso ao financeiro sem token JWT",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "periodo='mes' (sem token)",
            "procedimento": "1. Enviar GET /api/psicologos/financeiro sem header Authorization. "
                            "2. Verificar o código HTTP.",
            "resultado_esperado": "HTTP 401 (endpoint exige JWT)",
            "observacoes": "Acesso não autorizado",
        },
    ]


def _gerar_casos_websocket() -> list:
    """Casos de teste para a funcionalidade 'Chat em Tempo Real (WebSocket STOMP)'."""
    return [
        {
            "funcionalidade": "Chat em Tempo Real (WebSocket STOMP)",
            "objetivo": "Verificar conexão WebSocket com token JWT válido",
            "pre_condicao": "Paciente autenticado (JWT válido).",
            "dados_entrada": "ws://<host>/ws-chat?token=<jwt_válido>",
            "procedimento": "1. Obter token JWT do paciente demo. "
                            "2. Conectar ao endpoint /ws-chat?token=<token>. "
                            "3. Verificar se a conexão é estabelecida (CONNECTED).",
            "resultado_esperado": "Conexão estabelecida com sucesso (frame CONNECTED do STOMP)",
            "observacoes": "Cenário Positivo",
        },
        {
            "funcionalidade": "Chat em Tempo Real (WebSocket STOMP)",
            "objetivo": "Verificar rejeição de conexão WebSocket com token inválido",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "ws://<host>/ws-chat?token=token_invalido",
            "procedimento": "1. Conectar ao endpoint /ws-chat com token inválido. "
                            "2. Verificar se a conexão é recusada.",
            "resultado_esperado": "Conexão recusada (handshake falha — StompHandshakeInterceptor "
                                  "valida token na query string)",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Chat em Tempo Real (WebSocket STOMP)",
            "objetivo": "Verificar rejeição de conexão WebSocket com token expirado",
            "pre_condicao": "Token JWT expirado (24h de validade).",
            "dados_entrada": "ws://<host>/ws-chat?token=<jwt_expirado>",
            "procedimento": "1. Gerar token JWT expirado. "
                            "2. Conectar ao endpoint /ws-chat com token expirado. "
                            "3. Verificar se a conexão é recusada.",
            "resultado_esperado": "Conexão recusada (token expirado)",
            "observacoes": "Acesso não autorizado",
        },
        {
            "funcionalidade": "Chat em Tempo Real (WebSocket STOMP)",
            "objetivo": "Verificar envio e recebimento de mensagem via STOMP",
            "pre_condicao": "Paciente e psicólogo conectados ao WebSocket com tokens válidos.",
            "dados_entrada": "destination='/app/chat.send', body={\"destinatarioId\": <id_psicologo>, "
                             "\"mensagem\": \"Olá via STOMP\"}",
            "procedimento": "1. Conectar paciente e psicólogo ao WebSocket. "
                            "2. Paciente publica em /app/chat.send. "
                            "3. Verificar que o psicólogo recebe em /user/queue/mensagens.",
            "resultado_esperado": "Psicólogo recebe payload {\"type\": \"chat:message\", "
                                  "\"mensagem\": <Mensagem>}",
            "observacoes": "Cenário Positivo — chat em tempo real",
        },
        {
            "funcionalidade": "Chat em Tempo Real (WebSocket STOMP)",
            "objetivo": "Verificar conexão WebSocket sem token",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "ws://<host>/ws-chat (sem parâmetro token)",
            "procedimento": "1. Conectar ao endpoint /ws-chat sem token na query string. "
                            "2. Verificar se a conexão é recusada.",
            "resultado_esperado": "Conexão recusada (token ausente)",
            "observacoes": "Acesso não autorizado",
        },
    ]


def _gerar_casos_recuperar_senha() -> list:
    """Casos de teste para a funcionalidade 'Recuperação e Redefinição de Senha'."""
    return [
        {
            "funcionalidade": "Recuperação e Redefinição de Senha",
            "objetivo": "Verificar solicitação de recuperação de senha",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"email='{PACIENTE_DEMO}'",
            "procedimento": "1. Enviar POST /api/auth/recuperar-senha com email do paciente demo. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 200 com body {{\"message\": \"{MSG_RECUPERAR_SENHA}\"}}",
            "observacoes": "Cenário Positivo — mensagem genérica (não revela se email existe)",
        },
        {
            "funcionalidade": "Recuperação e Redefinição de Senha",
            "objetivo": "Verificar solicitação de recuperação com email inexistente",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": f"email='{_email_unico('naoexiste')}'",
            "procedimento": "1. Enviar POST /api/auth/recuperar-senha com email não cadastrado. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 200 com body {{\"message\": \"{MSG_RECUPERAR_SENHA}\"}} "
                                  "(mesma mensagem — não revela se email existe)",
            "observacoes": "Cenário Negativo — segurança por obscuridade",
        },
        {
            "funcionalidade": "Recuperação e Redefinição de Senha",
            "objetivo": "Verificar rejeição de recuperação com email vazio",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "email=''",
            "procedimento": "1. Enviar POST /api/auth/recuperar-senha com email vazio. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 400 com body {\"error\": \"Informe o email\"}",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Recuperação e Redefinição de Senha",
            "objetivo": "Verificar redefinição de senha com token inválido",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "token='token_invalido', novaSenha='NovaSenha@123'",
            "procedimento": "1. Enviar POST /api/auth/redefinir-senha com token inválido. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_TOKEN_INVALIDO}\"}}",
            "observacoes": "Cenário Negativo",
        },
        {
            "funcionalidade": "Recuperação e Redefinição de Senha",
            "objetivo": "Verificar rejeição de redefinição sem token",
            "pre_condicao": "Sistema no ar.",
            "dados_entrada": "novaSenha='NovaSenha@123' (sem token)",
            "procedimento": "1. Enviar POST /api/auth/redefinir-senha sem o campo token. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": "HTTP 400 com body {\"error\": \"Token e nova senha sao obrigatorios\"}",
            "observacoes": "Campo obrigatório vazio",
        },
        {
            "funcionalidade": "Recuperação e Redefinição de Senha",
            "objetivo": "Verificar rejeição de redefinição com token expirado",
            "pre_condicao": "Token de redefinição expirado (30 minutos de validade).",
            "dados_entrada": "token='<token_expirado>', novaSenha='NovaSenha@123'",
            "procedimento": "1. Enviar POST /api/auth/redefinir-senha com token expirado. "
                            "2. Verificar o código HTTP e o corpo da resposta.",
            "resultado_esperado": f"HTTP 400 com body {{\"error\": \"{MSG_TOKEN_INVALIDO}\"}}",
            "observacoes": "Valor no limite — expiração de 30 minutos",
        },
    ]


# ---------------------------------------------------------------------------
# Montagem da lista completa de casos de teste
# ---------------------------------------------------------------------------

def gerar_todos_casos() -> list:
    """
    Gera a lista completa de casos de teste para todas as funcionalidades.

    Returns:
        list: Lista de dicionários, cada um representando um caso de teste.
    """
    casos = []
    casos.extend(_gerar_casos_cadastro_paciente())
    casos.extend(_gerar_casos_cadastro_psicologo())
    casos.extend(_gerar_casos_login())
    casos.extend(_gerar_casos_agendamento())
    casos.extend(_gerar_casos_pagamento())
    casos.extend(_gerar_casos_chat())
    casos.extend(_gerar_casos_verificar_crp())
    casos.extend(_gerar_casos_alterar_senha())
    casos.extend(_gerar_casos_upload_foto())
    casos.extend(_gerar_casos_assinatura())
    casos.extend(_gerar_casos_usuarios())
    casos.extend(_gerar_casos_psicologos())
    casos.extend(_gerar_casos_websocket())
    casos.extend(_gerar_casos_recuperar_senha())
    return casos


# ---------------------------------------------------------------------------
# Geração da planilha Excel
# ---------------------------------------------------------------------------

# Colunas obrigatórias do modelo do PDF (item 8 do trabalho)
COLUNAS = [
    "Código",
    "Funcionalidade",
    "Objetivo",
    "Pré-condição",
    "Dados de entrada",
    "Procedimento",
    "Resultado esperado",
    "Resultado obtido",   # em branco — preenchimento manual
    "Status",             # em branco — preenchimento manual
    "Responsável",        # em branco — preenchimento manual
    "Data",               # em branco — preenchimento manual
    "Evidência",          # em branco — preenchimento manual
    "Observações",
]

# Larguras das colunas (em unidades de caracteres)
LARGURAS = {
    "Código": 10,
    "Funcionalidade": 32,
    "Objetivo": 45,
    "Pré-condição": 40,
    "Dados de entrada": 45,
    "Procedimento": 55,
    "Resultado esperado": 50,
    "Resultado obtido": 30,
    "Status": 12,
    "Responsável": 15,
    "Data": 12,
    "Evidência": 20,
    "Observações": 35,
}


def _estilizar_planilha(ws, num_linhas: int) -> None:
    """
    Aplica estilos (cabeçalho, bordas, alinhamento, quebra de linha) à planilha.

    Args:
        ws: Worksheet do openpyxl.
        num_linhas: Número total de linhas de dados.
    """
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    # Cabeçalho
    for col_idx, col_name in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = LARGURAS[col_name]

    # Corpo
    for row_idx in range(2, num_linhas + 2):
        for col_idx in range(1, len(COLUNAS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = body_alignment
            cell.border = thin_border

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Filtro automático
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{num_linhas + 1}"


def gerar_planilha(caminho: str | None = None) -> str:
    """
    Gera a planilha de casos de teste (.xlsx).

    Args:
        caminho: Caminho do arquivo de saída. Se None, usa
                 'dados/casos_de_teste.xlsx'.

    Returns:
        str: Caminho do arquivo gerado.
    """
    casos = gerar_todos_casos()

    if caminho is None:
        caminho = os.path.join(config.DADOS_DIR, "casos_de_teste.xlsx")

    os.makedirs(os.path.dirname(caminho), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Casos de Teste"

    # Preenche os dados
    for idx, caso in enumerate(casos, start=1):
        codigo = f"CT-{idx:03d}"
        ws.append([
            codigo,
            caso["funcionalidade"],
            caso["objetivo"],
            caso["pre_condicao"],
            caso["dados_entrada"],
            caso["procedimento"],
            caso["resultado_esperado"],
            "",  # Resultado obtido — em branco
            "",  # Status — em branco
            "",  # Responsável — em branco
            "",  # Data — em branco
            "",  # Evidência — em branco
            caso["observacoes"],
        ])

    _estilizar_planilha(ws, len(casos))
    wb.save(caminho)

    print(f"[OK] Planilha gerada: {caminho}")
    print(f"[OK] Total de casos de teste: {len(casos)}")
    return caminho


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    gerar_planilha()